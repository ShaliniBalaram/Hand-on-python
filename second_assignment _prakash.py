import pandas as pd

'''
import openpyxl
from openpyxl import load_workbook

excel_data = openpyxl.load_workbook("stock_data.xlsx")
sheet_obj = excel_data.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row

dataframe1 = pd.read_excel('stock_data.xlsx', header=0)
header = dataframe1.columns
sell_loc = dataframe1.columns.get_loc("BuyOrSell")
EP_loc = dataframe1.columns.get_loc("EntryPrice")
ExP_loc = dataframe1.columns.get_loc("ExitPrice")
lot_loc = dataframe1.columns.get_loc("Lot")

dataframe1.loc[dataframe1.iloc[:, sell_loc] == 'Sell', max_col+1] = dataframe1.iloc[:, ExP_loc] - dataframe1.iloc[:, EP_loc]
dataframe1.loc[dataframe1.iloc[:, sell_loc] == 'Sell', max_col+2] = dataframe1.loc[:, max_col+1] * dataframe1.iloc[:, lot_loc]
dataframe1.loc[dataframe1.iloc[:, sell_loc] == 'Buy', max_col+1] = dataframe1.iloc[:, EP_loc] - dataframe1.iloc[:, ExP_loc]
dataframe1.loc[dataframe1.iloc[:, sell_loc] == 'Buy', max_col+2] = dataframe1.loc[:, max_col+1] * dataframe1.iloc[:, lot_loc]

dataframe1.update

writer = pd.ExcelWriter('stock_data.xlsx', engine='openpyxl', mode='a')
writer.book = openpyxl.load_workbook('stock_data.xlsx')
dataframe1.to_excel(writer, sheet_name='dataframe1')
writer.save()
writer.close()
'''

df = pd.read_excel('stock_data.xlsx')

print(len(df))

# create new columns with default value as 0
df['GrossValue'] = 0

df.loc[(df['BuyOrSell'] == 'Sell'), 'GrossValue'] = df['ExitPrice'] - df['EntryPrice']
df.loc[(df['BuyOrSell'] == 'Buy'), 'GrossValue'] = df['EntryPrice'] - df['ExitPrice']

df['PNL'] = df['GrossValue'] * df['Lot']

# show columns
print(df.columns)

# show top 5 rows
print(df.head(5))


# write to file
df.to_excel('stock_data_output.xlsx', index=False)
