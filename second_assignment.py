import pandas as pd
import openpyxl
from openpyxl import load_workbook

excel_data = openpyxl.load_workbook("stock_data.xlsx")
sheet_obj = excel_data.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row
dataframe1 = pd.read_excel('stock_data.xlsx', header=0)
sell_loc = dataframe1.columns.get_loc("BuyOrSell")
EP_loc = dataframe1.columns.get_loc("EntryPrice")
ExP_loc = dataframe1.columns.get_loc("ExitPrice")
lot_loc = dataframe1.columns.get_loc("Lot")

dataframe2 = dataframe1

dataframe2.loc[dataframe1.iloc[:, sell_loc] == 'Sell', max_col+1] = dataframe2.iloc[:, ExP_loc]-dataframe2.iloc[:, EP_loc]
dataframe2.loc[dataframe1.iloc[:, sell_loc] == 'Buy', max_col+1] = dataframe2.iloc[:, EP_loc]-dataframe2.iloc[:, ExP_loc]
dataframe2.loc[:, max_col+2] = dataframe2.loc[:, max_col+1] * dataframe2.iloc[:, lot_loc]
dataframe2.update
dataframe1["GrossValue"] = dataframe2.loc[:, max_col+1]
dataframe1["PNL"] = dataframe2.loc[:, max_col+2]

dataframe1.to_excel("stock_data.xlsx")

